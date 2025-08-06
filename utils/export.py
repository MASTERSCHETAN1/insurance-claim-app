"""
Export utilities for insurance claim data
"""

import csv
import os
from typing import List, Dict
from datetime import datetime

# Try to import openpyxl for Excel support
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

class ExportManager:
    """Manager for exporting claim data to various formats"""
    
    def __init__(self):
        # Define export columns and their display names
        self.export_columns = [
            ('id', 'Claim ID'),
            ('entry_date', 'Entry Date'),
            ('admission_date', 'Admission Date'),
            ('customer_name', 'Customer Name'),
            ('policy_number', 'Policy Number'),
            ('hospital_name', 'Hospital Name'),
            ('company_name', 'Company Name'),
            ('claim_number', 'Claim Number'),
            ('claim_status', 'Claim Status'),
            ('claimed_amount', 'Claimed Amount'),
            ('approved_amount', 'Approved Amount'),
            ('claim_type', 'Claim Type'),
            ('remark', 'Remark'),
            ('parent_claim_id', 'Parent Claim ID'),
            ('created_at', 'Created At'),
            ('updated_at', 'Updated At')
        ]
    
    def export_to_csv(self, claims: List[Dict], filename: str) -> bool:
        """
        Export claims to CSV file
        
        Args:
            claims: List of claim dictionaries
            filename: Output filename
            
        Returns:
            True if successful, False otherwise
        """
        try:
            with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                
                # Write headers
                headers = [display_name for _, display_name in self.export_columns]
                writer.writerow(headers)
                
                # Write claim data
                for claim in claims:
                    row = []
                    for field_name, _ in self.export_columns:
                        value = claim.get(field_name, '')
                        
                        # Format specific fields
                        if field_name in ['claimed_amount', 'approved_amount'] and value:
                            value = f"{value:.2f}"
                        elif field_name in ['created_at', 'updated_at'] and value:
                            try:
                                # Format timestamp
                                if isinstance(value, str):
                                    dt = datetime.fromisoformat(value.replace('Z', '+00:00'))
                                    value = dt.strftime('%Y-%m-%d %H:%M:%S')
                            except:
                                pass
                        elif value is None:
                            value = ''
                        
                        row.append(str(value))
                    
                    writer.writerow(row)
            
            return True
            
        except Exception as e:
            print(f"Error exporting to CSV: {e}")
            return False
    
    def export_to_excel(self, claims: List[Dict], filename: str) -> bool:
        """
        Export claims to Excel file
        
        Args:
            claims: List of claim dictionaries
            filename: Output filename
            
        Returns:
            True if successful, False otherwise
        """
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl library not available. Please install it to export to Excel.")
        
        try:
            # Create workbook and worksheet
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Insurance Claims"
            
            # Header styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Write headers
            headers = [display_name for _, display_name in self.export_columns]
            for col_idx, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Write claim data
            for row_idx, claim in enumerate(claims, 2):
                for col_idx, (field_name, _) in enumerate(self.export_columns, 1):
                    value = claim.get(field_name, '')
                    
                    # Format specific fields
                    if field_name in ['claimed_amount', 'approved_amount'] and value:
                        # Keep as number for Excel calculations
                        cell = worksheet.cell(row=row_idx, column=col_idx, value=float(value))
                        cell.number_format = '₹#,##0.00'
                    elif field_name in ['created_at', 'updated_at'] and value:
                        try:
                            # Format timestamp
                            if isinstance(value, str):
                                dt = datetime.fromisoformat(value.replace('Z', '+00:00'))
                                cell = worksheet.cell(row=row_idx, column=col_idx, value=dt)
                                cell.number_format = 'YYYY-MM-DD HH:MM:SS'
                            else:
                                cell = worksheet.cell(row=row_idx, column=col_idx, value=str(value))
                        except:
                            cell = worksheet.cell(row=row_idx, column=col_idx, value=str(value))
                    elif field_name in ['entry_date', 'admission_date'] and value:
                        try:
                            # Format dates
                            dt = datetime.strptime(value, '%Y-%m-%d')
                            cell = worksheet.cell(row=row_idx, column=col_idx, value=dt)
                            cell.number_format = 'YYYY-MM-DD'
                        except:
                            cell = worksheet.cell(row=row_idx, column=col_idx, value=str(value))
                    else:
                        cell = worksheet.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else '')
            
            # Auto-adjust column widths
            for col_idx in range(1, len(headers) + 1):
                column_letter = get_column_letter(col_idx)
                column_cells = [worksheet[f"{column_letter}{row}"] for row in range(1, len(claims) + 2)]
                max_length = max(len(str(cell.value or '')) for cell in column_cells)
                worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
            
            # Add summary information
            summary_row = len(claims) + 3
            worksheet.cell(row=summary_row, column=1, value="Summary").font = Font(bold=True)
            
            # Total claims
            worksheet.cell(row=summary_row + 1, column=1, value="Total Claims:")
            worksheet.cell(row=summary_row + 1, column=2, value=len(claims))
            
            # Total claimed amount
            claimed_col = next(i for i, (field, _) in enumerate(self.export_columns, 1) if field == 'claimed_amount')
            worksheet.cell(row=summary_row + 2, column=1, value="Total Claimed Amount:")
            worksheet.cell(row=summary_row + 2, column=2, value=f"=SUM({get_column_letter(claimed_col)}2:{get_column_letter(claimed_col)}{len(claims) + 1})")
            worksheet.cell(row=summary_row + 2, column=2).number_format = '₹#,##0.00'
            
            # Total approved amount
            approved_col = next(i for i, (field, _) in enumerate(self.export_columns, 1) if field == 'approved_amount')
            worksheet.cell(row=summary_row + 3, column=1, value="Total Approved Amount:")
            worksheet.cell(row=summary_row + 3, column=2, value=f"=SUM({get_column_letter(approved_col)}2:{get_column_letter(approved_col)}{len(claims) + 1})")
            worksheet.cell(row=summary_row + 3, column=2).number_format = '₹#,##0.00'
            
            # Export date
            worksheet.cell(row=summary_row + 5, column=1, value="Exported on:")
            worksheet.cell(row=summary_row + 5, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            
            # Save workbook
            workbook.save(filename)
            return True
            
        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            return False
    
    def export_claims_by_status(self, claims: List[Dict], filename: str, format_type: str = 'excel') -> bool:
        """
        Export claims grouped by status
        
        Args:
            claims: List of claim dictionaries
            filename: Output filename
            format_type: 'excel' or 'csv'
            
        Returns:
            True if successful, False otherwise
        """
        # Group claims by status
        claims_by_status = {}
        for claim in claims:
            status = claim.get('claim_status', 'Unknown')
            if status not in claims_by_status:
                claims_by_status[status] = []
            claims_by_status[status].append(claim)
        
        if format_type.lower() == 'excel' and EXCEL_AVAILABLE:
            return self._export_grouped_excel(claims_by_status, filename)
        else:
            return self._export_grouped_csv(claims_by_status, filename)
    
    def _export_grouped_excel(self, claims_by_status: Dict[str, List[Dict]], filename: str) -> bool:
        """Export grouped claims to Excel with multiple sheets"""
        try:
            workbook = openpyxl.Workbook()
            
            # Remove default sheet
            default_sheet = workbook.active
            
            for status, status_claims in claims_by_status.items():
                # Create sheet for each status
                worksheet = workbook.create_sheet(title=status[:31])  # Sheet name limit
                
                # Header styling
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                # Write headers
                headers = [display_name for _, display_name in self.export_columns]
                for col_idx, header in enumerate(headers, 1):
                    cell = worksheet.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                
                # Write claim data
                for row_idx, claim in enumerate(status_claims, 2):
                    for col_idx, (field_name, _) in enumerate(self.export_columns, 1):
                        value = claim.get(field_name, '')
                        if field_name in ['claimed_amount', 'approved_amount'] and value:
                            cell = worksheet.cell(row=row_idx, column=col_idx, value=float(value))
                            cell.number_format = '₹#,##0.00'
                        else:
                            worksheet.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else '')
                
                # Auto-adjust column widths
                for col_idx in range(1, len(headers) + 1):
                    column_letter = get_column_letter(col_idx)
                    worksheet.column_dimensions[column_letter].width = 15
            
            # Remove default sheet if it's empty
            if default_sheet.title == "Sheet" and len(workbook.sheetnames) > 1:
                workbook.remove(default_sheet)
            
            workbook.save(filename)
            return True
            
        except Exception as e:
            print(f"Error exporting grouped Excel: {e}")
            return False
    
    def _export_grouped_csv(self, claims_by_status: Dict[str, List[Dict]], filename: str) -> bool:
        """Export grouped claims to CSV with status sections"""
        try:
            with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                
                # Write summary header
                writer.writerow(['Insurance Claims Export by Status'])
                writer.writerow(['Exported on:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
                writer.writerow([])  # Empty row
                
                for status, status_claims in claims_by_status.items():
                    # Status header
                    writer.writerow([f'Status: {status} ({len(status_claims)} claims)'])
                    writer.writerow([])
                    
                    # Column headers
                    headers = [display_name for _, display_name in self.export_columns]
                    writer.writerow(headers)
                    
                    # Claim data
                    for claim in status_claims:
                        row = []
                        for field_name, _ in self.export_columns:
                            value = claim.get(field_name, '')
                            if field_name in ['claimed_amount', 'approved_amount'] and value:
                                value = f"{value:.2f}"
                            elif value is None:
                                value = ''
                            row.append(str(value))
                        writer.writerow(row)
                    
                    writer.writerow([])  # Empty row between status groups
            
            return True
            
        except Exception as e:
            print(f"Error exporting grouped CSV: {e}")
            return False
    
    def get_export_formats(self) -> List[str]:
        """Get list of available export formats"""
        formats = ['CSV']
        if EXCEL_AVAILABLE:
            formats.append('Excel')
        return formats
